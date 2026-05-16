import openpyxl
from collections import Counter

FILE = "1 KK_ART Pondokrejo.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

# Ambil semua baris
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    rows.append(row)

# Header di baris index 2 (Row 3)
headers = rows[2]

# Kolom yang kita cek (bansos)
# IS Bpnt=105, IS Pkh=108, IS Blt=111, IS Listrik=114, IS Banpem=117, IS Lpg=123, IS Baznas=126
cols_to_check = {
    105: "IS Bpnt",
    107: "Bpnt Thn",
    108: "IS Pkh",
    110: "Pkh Thn",
    111: "IS Blt",
    114: "IS Listrik",
    123: "IS Lpg",
    126: "IS Baznas",
    51: "IS Ekstrem",
    52: "IS Stun",
}

print("=== VERIFIKASI KOLOM BANSOS (dari file asli) ===")
print(f"Total baris data: {len(rows) - 3}")
print()

# Untuk setiap kolom, hitung distribusi nilai
for col_idx, col_name in cols_to_check.items():
    actual_header = headers[col_idx] if col_idx < len(headers) else "N/A"
    values = []
    for row in rows[3:]:  # skip 3 header rows
        if col_idx < len(row) and row[col_idx] is not None:
            values.append(str(row[col_idx]).strip())
    
    counter = Counter(values)
    total_1 = counter.get("1", 0)
    total_2 = counter.get("2", 0)
    total_empty = len(rows) - 3 - len(values)
    
    print(f"[{col_idx:>3}] {col_name:15} | Header: '{actual_header}' | Nilai '1'={total_1}, '2'={total_2}, kosong={total_empty}")
    # Show top 5 values
    for val, cnt in counter.most_common(5):
        print(f"       -> '{val}': {cnt}x")
    print()

wb.close()
