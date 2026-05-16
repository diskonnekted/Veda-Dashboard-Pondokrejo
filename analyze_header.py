import openpyxl

FILE = "1 KK_ART Pondokrejo.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
sheets = wb.sheetnames

print(f"Sheet: {sheets}")

ws = wb[sheets[0]]

# Ambil 3 baris pertama untuk memahami struktur
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    rows.append(row)
    if i >= 5:
        break

print(f"\n=== BARIS 0 (Row 1) ===")
for idx, v in enumerate(rows[0]):
    if v is not None and str(v).strip() != '':
        print(f"[{idx:>3}] {str(v)[:80]}")

print(f"\n=== BARIS 1 (Row 2 - kode variabel) ===")
for idx, v in enumerate(rows[1]):
    if v is not None and str(v).strip() != '':
        print(f"[{idx:>3}] {str(v)[:80]}")

print(f"\n=== BARIS 2 (Row 3 - sub-header?) ===")
for idx, v in enumerate(rows[2]):
    if v is not None and str(v).strip() != '':
        print(f"[{idx:>3}] {str(v)[:80]}")

print(f"\n=== BARIS 3 (Row 4 - data pertama?) ===")
for idx, v in enumerate(rows[3]):
    if v is not None and str(v).strip() != '':
        print(f"[{idx:>3}] {str(v)[:80]}")

wb.close()
