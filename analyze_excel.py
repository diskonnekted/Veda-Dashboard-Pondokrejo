import openpyxl
import json
from collections import defaultdict

FILE = "1 KK_ART Pondokrejo.xlsx"
print(f"Membuka: {FILE}")

wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
sheets = wb.sheetnames
print(f"\nJumlah Sheet: {len(sheets)}")
print(f"Nama Sheet: {sheets}")

results = {}

for sheet_name in sheets:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")

    # Ambil semua baris, maks 5 baris pertama untuk header + sample
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 4:  # header + 4 baris data
            break

    if not rows:
        print("  [KOSONG]")
        continue

    # Baris pertama = header
    headers = [str(c) if c is not None else f"COL_{i}" for i, c in enumerate(rows[0])]
    print(f"\nJumlah Kolom: {len(headers)}")
    print(f"\nDaftar Kolom (index: nama):")
    for idx, h in enumerate(headers):
        print(f"  [{idx:>3}] {h}")

    # Sample data dari baris ke-2
    if len(rows) > 1:
        print(f"\nSample Baris 2:")
        for idx, (h, v) in enumerate(zip(headers, rows[1])):
            if v is not None and str(v).strip() != '':
                print(f"  [{idx:>3}] {h} = {str(v)[:60]}")

    # Hitung baris total (read_only compat)
    print(f"\nDimensions (max_row): {ws.max_row}")
    results[sheet_name] = {
        "columns": headers,
        "max_row": ws.max_row
    }

wb.close()

# Simpan hasil ke JSON
with open("excel_analysis.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print("\n\n=== ANALISIS SELESAI. Hasil disimpan ke excel_analysis.json ===")
