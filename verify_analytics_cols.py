import openpyxl
from collections import Counter

FILE = "1 KK_ART Pondokrejo.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = []
for row in ws.iter_rows(values_only=True):
    rows.append(row)

headers = rows[2]  # Row 3 = column names

# Verify key columns
cols = {
    # Household level
    71:  "ID Fasbab (Sanitasi)",
    59:  "ID Airminum (Air)",
    29:  "Col 29 (House Status?)",
    55:  "Lantai Luas",
    56:  "ID Lantai",
    57:  "ID Dinding",
    38:  "ID Desil",
    # Member level
    234: "ID Hub Keluarga",
    237: "ID Kelamin",
    252: "ID Kerja",
    245: "ID Jenjang",
    260: "Usia",
    242: "Hamil",
}

print("=== VERIFIKASI KOLOM KUNCI ===\n")
for col_idx, desc in cols.items():
    actual_hdr = headers[col_idx] if col_idx < len(headers) else "N/A"
    values = []
    for row in rows[3:]:
        if col_idx < len(row) and row[col_idx] is not None:
            values.append(str(row[col_idx]).strip())
    counter = Counter(values)
    print(f"[{col_idx:>3}] {desc} | Header='{actual_hdr}'")
    for v, c in counter.most_common(8):
        print(f"      '{v}': {c}x")
    print()

wb.close()
