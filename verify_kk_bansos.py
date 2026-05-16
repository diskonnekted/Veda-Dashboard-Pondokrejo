import openpyxl
from collections import Counter, defaultdict

FILE = "1 KK_ART Pondokrejo.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = []
for row in ws.iter_rows(values_only=True):
    rows.append(row)

# Headers at row index 2
headers = rows[2]

# Data starts at row index 3
# NO KK = col 2, Nama KK = col 14 (per header baris ke-2)
# Let's verify
print("Col 1 header:", headers[1])   # ID Ruta
print("Col 2 header:", headers[2])   # NO KK
print("Col 14 header:", headers[14]) # Nama Kepala KK

# Check NO KK uniqueness
kk_set = defaultdict(list)
kk_bpnt = defaultdict(set)

for i, row in enumerate(rows[3:], start=3):
    if len(row) < 15:
        continue
    nokk = str(row[2]).strip() if row[2] else ""
    headname = str(row[14]).strip() if row[14] else ""
    is_bpnt = str(row[105]).strip() if len(row) > 105 and row[105] is not None else ""
    
    if nokk:
        unique_key = nokk + "|" + headname
        kk_set[unique_key].append(is_bpnt)

# Count unique KK with BPNT
bpnt_kk = 0
pkh_kk = 0
blt_kk = 0

kk_bpnt_check = defaultdict(list)
kk_pkh_check = defaultdict(list)
kk_blt_check = defaultdict(list)

for i, row in enumerate(rows[3:], start=3):
    if len(row) < 15:
        continue
    nokk = str(row[2]).strip() if row[2] else ""
    headname = str(row[14]).strip() if row[14] else ""
    
    if not nokk:
        continue
    
    unique_key = nokk + "|" + headname
    
    bpnt = str(row[105]).strip() if len(row) > 105 and row[105] is not None else ""
    pkh  = str(row[108]).strip() if len(row) > 108 and row[108] is not None else ""
    blt  = str(row[111]).strip() if len(row) > 111 and row[111] is not None else ""
    
    if bpnt:
        kk_bpnt_check[unique_key].append(bpnt)
    if pkh:
        kk_pkh_check[unique_key].append(pkh)
    if blt:
        kk_blt_check[unique_key].append(blt)

# Count unique KK yang punya "1"
bpnt_unique = sum(1 for k, vals in kk_bpnt_check.items() if "1" in vals)
pkh_unique  = sum(1 for k, vals in kk_pkh_check.items() if "1" in vals)
blt_unique  = sum(1 for k, vals in kk_blt_check.items() if "1" in vals)

total_kk = len(kk_set)

print(f"\nTotal KK unik: {total_kk}")
print(f"\nKK penerima BPNT: {bpnt_unique}")
print(f"KK penerima PKH : {pkh_unique}")
print(f"KK penerima BLT : {blt_unique}")

# Cek sample KK dengan ART terbanyak
sample = sorted(kk_set.items(), key=lambda x: -len(x[1]))[:3]
print("\nSample KK dengan banyak baris ART:")
for k, v in sample:
    print(f"  {k[:40]}: {len(v)} baris, IS_Bpnt values: {set(v)}")

wb.close()
